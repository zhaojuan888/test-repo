# -*- coding: utf-8 -*-
"""
新列报准则（财会〔2026〕11号）利润表五分类重排模板生成器
- 公式驱动：输出区全部用 Excel 公式引用输入区，改任一输入全套自动重排
- 五分类：经营 / 投资 / 筹资 / 所得税费用 / 终止经营（准则第三十二条）
- 小计级联：经营利润 → 经营及投资利润 → 持续经营利润(税前) → 持续经营净利润 → 净利润
- 关键重分类：利息收入→投资、利息费用→筹资、汇兑损益跟项目、公允价值变动按类归集、终止经营隔离
- 含「新旧净利润勾稽校验」

用法：python build_pnl_fivecat_template.py
输出：../9-产出/新列报准则利润表五分类重排模板.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"D:\AI学习\审迹知识库\9-产出\新列报准则利润表五分类重排模板.xlsx"
IN_SHEET = "①旧格式-输入"
OUT_SHEET = "②新五分类-输出"
MAP_SHEET = "③映射对照"
CHK_SHEET = "④2027切换清单"

# ---------- 样式 ----------
TITLE = Font(name="微软雅黑", size=14, bold=True, color="1F3864")
HDR = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="微软雅黑", size=10, bold=True)
NORM = Font(name="微软雅黑", size=10)
SUB = Font(name="微软雅黑", size=10, bold=True, color="1F3864")
SMALL = Font(name="微软雅黑", size=9, color="808080")
HDR_FILL = PatternFill("solid", fgColor="1F3864")
CAT_FILL = PatternFill("solid", fgColor="D6DCE4")
SUB_FILL = PatternFill("solid", fgColor="EAF0F7")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREEN = PatternFill("solid", fgColor="E2EFDA")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CEN; cell.border = BORDER

def box(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER

# ============================================================
# 输入区布局（①旧格式-输入）
# 行号固定，供输出区引用。B列=符号(+1/-1)，C=本期，D=上期
# ============================================================
# (行项目, 符号)  —— 行号 = index+3
INPUT_ROWS = [
    ("营业收入", 1),
    ("减：营业成本", -1),
    ("税金及附加", -1),
    ("销售费用", -1),
    ("管理费用", -1),
    ("研发费用", -1),
    ("财务费用——经营部分（手续费/银行费等）", -1),
    ("  其中：利息费用（→筹资类）", -1),
    ("  其中：利息收入（→投资类）", 1),
    ("其他收益", 1),
    ("投资收益", 1),
    ("公允价值变动收益（按类别归集，默认投资）", 1),
    ("信用减值损失", -1),
    ("资产减值损失", -1),
    ("资产处置收益", 1),
    ("营业外收入（默认经营类兜底）", 1),
    ("减：营业外支出（默认经营类兜底）", -1),
    ("汇兑损益——经营（按来源填正负）", 0),
    ("汇兑损益——投资（按来源填正负）", 0),
    ("汇兑损益——筹资（按来源填正负）", 0),
    ("所得税费用", -1),
    ("终止经营损益（税后净额，填正负）", 0),
]
# 行号映射：input row index -> excel row (起始3)
def IN(row_idx, col):
    """返回输入区某行的单元格引用串，row_idx 为 INPUT_ROWS 下标(0起)"""
    r = row_idx + 3
    return f"'{IN_SHEET}'!{get_column_letter(col)}{r}"

wb = openpyxl.Workbook()

# ---------------- ① 旧格式-输入 ----------------
ws = wb.active
ws.title = IN_SHEET
ws.merge_cells("A1:D1")
ws["A1"] = "利润表（旧格式 / 现行列报）—— 数据输入区"
ws["A1"].font = TITLE
# 表头 row2（数据从 row3 开始，行号固定供输出区引用）
for c, h in enumerate(["行项目", "符号", "本期金额", "上期金额"], start=1):
    ws.cell(row=2, column=c, value=h)
style_header(ws, 2, 4)

for i, (name, sign) in enumerate(INPUT_ROWS):
    r = i + 3
    ws.cell(row=r, column=1, value=name).font = NORM
    ws.cell(row=r, column=2, value=sign).font = NORM
    ws.cell(row=r, column=2).alignment = CEN
    ws.cell(row=r, column=3).fill = YELLOW   # 本期
    ws.cell(row=r, column=4).fill = YELLOW   # 上期
    ws.cell(row=r, column=3).number_format = '#,##0.00'
    ws.cell(row=r, column=4).number_format = '#,##0.00'
    ws.cell(row=r, column=3).alignment = RIGHT
    ws.cell(row=r, column=4).alignment = RIGHT

ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 6
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 16
box(ws, 2, 1, 2 + len(INPUT_ROWS), 4)
nr = 2 + len(INPUT_ROWS) + 1
ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=4)
ws.cell(row=nr, column=1, value="黄底单元格为手工录入（或接 TB/试算平衡表取数）；符号列驱动输出区正负。留空按 0 处理。").font = SMALL
ws.freeze_panes = "A3"

# ---------------- ② 新五分类-输出 ----------------
ws2 = wb.create_sheet(OUT_SHEET)
ws2.merge_cells("A1:D1")
ws2["A1"] = "利润表（新列报准则·五分类重排）—— 公式输出区"
ws2["A1"].font = TITLE
for c, h in enumerate(["类别", "行项目", "本期金额", "上期金额"], start=1):
    ws2.cell(row=2, column=c, value=h)
style_header(ws2, 2, 4)

r = 3
def put(cat, name, formula, is_sub=False, is_cat=False):
    global r
    ws2.cell(row=r, column=1, value=cat)
    ws2.cell(row=r, column=2, value=name)
    ws2.cell(row=r, column=3, value=formula)
    ws2.cell(row=r, column=4, value=formula.replace("C", "D") if formula else None)
    for c in (1, 2, 3, 4):
        cell = ws2.cell(row=r, column=c)
        cell.border = BORDER
        if is_cat:
            cell.fill = CAT_FILL; cell.font = BOLD
        elif is_sub:
            cell.fill = SUB_FILL; cell.font = SUB
        else:
            cell.font = NORM
    ws2.cell(row=r, column=3).number_format = '#,##0.00'
    ws2.cell(row=r, column=4).number_format = '#,##0.00'
    ws2.cell(row=r, column=3).alignment = RIGHT
    ws2.cell(row=r, column=4).alignment = RIGHT
    rr = r
    r += 1
    return rr

# 经营类
put("【经营类】", "营业收入", f"={IN(0,3)}")
put("", "减：营业成本", f"={IN(1,3)}*{INPUT_ROWS[1][1]}")
put("", "税金及附加", f"={IN(2,3)}*{INPUT_ROWS[2][1]}")
put("", "销售费用", f"={IN(3,3)}*{INPUT_ROWS[3][1]}")
put("", "管理费用", f"={IN(4,3)}*{INPUT_ROWS[4][1]}")
put("", "研发费用", f"={IN(5,3)}*{INPUT_ROWS[5][1]}")
put("", "财务费用——经营部分", f"={IN(6,3)}*{INPUT_ROWS[6][1]}")
put("", "信用减值损失", f"={IN(12,3)}*{INPUT_ROWS[12][1]}")
put("", "资产减值损失", f"={IN(13,3)}*{INPUT_ROWS[13][1]}")
put("", "资产处置收益", f"={IN(14,3)}*{INPUT_ROWS[14][1]}")
put("", "其他收益", f"={IN(9,3)}*{INPUT_ROWS[9][1]}")
put("", "营业外收入（兜底）", f"={IN(15,3)}*{INPUT_ROWS[15][1]}")
put("", "减：营业外支出（兜底）", f"={IN(16,3)}*{INPUT_ROWS[16][1]}")
put("", "汇兑损益——经营", f"={IN(17,3)}")
op_profit = put("【经营类】", "经营利润（小计）", f"=SUM(C3:C{r-1})", is_sub=True)

# 投资类
put("【投资类】", "投资收益", f"={IN(10,3)}*{INPUT_ROWS[10][1]}")
put("", "利息收入", f"={IN(8,3)}*{INPUT_ROWS[8][1]}")
put("", "公允价值变动收益（归投资）", f"={IN(11,3)}*{INPUT_ROWS[11][1]}")
put("", "汇兑损益——投资", f"={IN(18,3)}")
inv_sum = put("【投资类】", "投资类小计", f"=SUM(C{op_profit+1}:C{r-1})", is_sub=True)
oi_profit = put("【经营+投资】", "经营及投资利润", f"=C{op_profit}+C{inv_sum}", is_sub=True)

# 筹资类
fin_start = r
put("【筹资类】", "利息费用", f"={IN(7,3)}*{INPUT_ROWS[7][1]}")
put("", "汇兑损益——筹资", f"={IN(19,3)}")
fin_sum = put("【筹资类】", "筹资类小计", f"=SUM(C{fin_start}:C{r-1})", is_sub=True)
cont_pre = put("【持续经营·税前】", "持续经营利润（税前）", f"=C{oi_profit}-C{fin_sum}", is_sub=True)

# 所得税费用类
put("【所得税费用类】", "所得税费用", f"={IN(20,3)}*{INPUT_ROWS[20][1]}")
cont_net = put("【持续经营·税后】", "持续经营净利润", f"=C{cont_pre}-C{r-1}", is_sub=True)

# 终止经营类
put("【终止经营类】", "终止经营损益（税后）", f"={IN(21,3)}")
net = put("【净利润】", "净利润", f"=C{cont_net}+C{r-1}", is_sub=True)

# 校验区
r += 1
ws2.cell(row=r, column=1, value="校验").font = BOLD
ws2.cell(row=r, column=2, value="旧表净利润（按旧格式重算）").font = NORM
# 旧格式净利润 = 所有带符号项之和
old_terms = []
for i, (nm, sg) in enumerate(INPUT_ROWS):
    ref = f"{IN(i,3)}"
    if sg == 1:
        old_terms.append(ref)
    elif sg == -1:
        old_terms.append(f"-{ref}")
    else:
        old_terms.append(ref)  # 汇兑/终止经营 已带正负
old_formula = "=" + "+".join(old_terms)
ws2.cell(row=r, column=3, value=old_formula).number_format = '#,##0.00'
old_net_row = r
r += 1
ws2.cell(row=r, column=2, value="新表净利润（输出区）").font = NORM
ws2.cell(row=r, column=3, value=f"=C{net}").number_format = '#,##0.00'
new_net_row = r
r += 1
ws2.cell(row=r, column=2, value="差异（应为 0）").font = BOLD
diff_cell = ws2.cell(row=r, column=3, value=f"=C{new_net_row}-C{old_net_row}")
diff_cell.number_format = '#,##0.00'; diff_cell.font = BOLD; diff_cell.fill = GREEN
ws2.cell(row=r, column=4, value="若≠0，检查分类或符号录入").font = SMALL

r += 2
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws2.cell(row=r, column=1, value="本表全部为公式，引用「①旧格式-输入」。改输入→本表自动重排。小计级联：经营利润→经营及投资利润→持续经营利润(税前)→持续经营净利润→净利润。").font = SMALL

ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 34
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 16
ws2.freeze_panes = "A3"

# ---------------- ③ 映射对照 ----------------
ws3 = wb.create_sheet(MAP_SHEET)
ws3.merge_cells("A1:C1")
ws3["A1"] = "旧格式行项目 → 五分类 映射对照（准则依据）"
ws3["A1"].font = TITLE
for c, h in enumerate(["旧格式行项目", "五分类归属", "准则依据 / 备注"], start=1):
    ws3.cell(row=2, column=c, value=h)
style_header(ws3, 2, 3)

MAP = [
    ("营业收入", "经营类", "第三十二条：经营类为剩余兜底类别"),
    ("减：营业成本", "经营类", "主业成本"),
    ("税金及附加", "经营类", "经营相关税费"),
    ("销售费用", "经营类", "经营类"),
    ("管理费用", "经营类", "经营类"),
    ("研发费用", "经营类", "经营类"),
    ("财务费用——经营部分", "经营类", "手续费等，非利息收支"),
    ("其中：利息费用", "筹资类", "第三十四条：利息费用归入筹资类别"),
    ("其中：利息收入", "投资类", "第三十四条：利息收入归入投资类别（特定资产回报）"),
    ("其他收益", "经营类", "经营相关政府补助等"),
    ("投资收益", "投资类", "对联营合营、债权投资等回报"),
    ("公允价值变动收益", "投资类（默认）", "取消单独项目，按所属类别归集；投资类为主"),
    ("信用减值损失", "经营类", "经营类（除非对应投资资产）"),
    ("资产减值损失", "经营类", "经营类（除非对应投资资产）"),
    ("资产处置收益", "经营类", "经营相关资产处置"),
    ("营业外收入", "经营类（兜底）", "★判断点：非终止经营的营业外收支默认归经营；属终止经营应剔除"),
    ("减：营业外支出", "经营类（兜底）", "★同上，注意与终止经营区分"),
    ("汇兑损益——经营", "经营类", "跟项目走：应收/应付等经营项目产生→经营"),
    ("汇兑损益——投资", "投资类", "跟项目走：外币存款/投资产生→投资"),
    ("汇兑损益——筹资", "筹资类", "跟项目走：外币借款产生→筹资"),
    ("所得税费用", "所得税费用类", "单独列示；持续/终止可拆分"),
    ("终止经营损益", "终止经营类", "独立隔离列示，不混入持续经营"),
]
for i, (a, b, c) in enumerate(MAP):
    rr = i + 3
    ws3.cell(row=rr, column=1, value=a).font = NORM
    ws3.cell(row=rr, column=2, value=b).font = NORM
    ws3.cell(row=rr, column=3, value=c).font = NORM
    ws3.cell(row=rr, column=3).alignment = LEFT
box(ws3, 2, 1, 2 + len(MAP), 3)
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 52

# ---------------- ④ 2027切换清单 ----------------
ws4 = wb.create_sheet(CHK_SHEET)
ws4.merge_cells("A1:E1")
ws4["A1"] = "新列报准则（财会〔2026〕11号）2027 切换 Checklist"
ws4["A1"].font = TITLE
ws4.merge_cells("A2:E2")
ws4["A2"] = "A+H / 境外上市：2027-01-01 起执行且需追溯重述对比期（2025、2026）。纯境内上市 2029；非上市 2030。允许提前。"
ws4["A2"].font = SMALL
for c, h in enumerate(["序号", "切换事项", "是否完成", "责任人", "备注"], start=1):
    ws4.cell(row=3, column=c, value=h)
style_header(ws4, 3, 5)

CHECKS = [
    ("梳理会计科目体系，确保科目设置与新利润表五分类衔接", ""),
    ("识别「特定主要业务活动」主体（银行/地产/融资租赁等），其投资/筹资损益归入经营类", "★关键点"),
    ("拆分财务费用为：利息收入(投资)/利息费用(筹资)/其他(经营)", ""),
    ("汇兑损益按来源拆分至 经营/投资/筹资（跟项目走，勿一把抓进财务费用）", "★审计必查"),
    ("公允价值变动损益取消单独项目，按所属类别归集", ""),
    ("终止经营损益独立隔离列示，不混入持续经营", ""),
    ("追溯重述对比期间（2025、2026）财务数据至新五分类格式", "★2027强制"),
    ("识别并披露管理层业绩指标(MPM)，附注说明内涵与计算方式（第六十六条）", "★突破点"),
    ("评估信息系统 / 报表模块改造，支持五分类取数与小计级联", ""),
    ("关键判断（特定主要业务活动、MPM界定）形成书面记录备查", ""),
    ("财政部应用指南 / 会计科目修订出台后，复核并调整分类口径", ""),
]
for i, (item, note) in enumerate(CHECKS):
    rr = i + 4
    ws4.cell(row=rr, column=1, value=i + 1).alignment = CEN
    ws4.cell(row=rr, column=2, value=item).font = NORM
    ws4.cell(row=rr, column=2).alignment = LEFT
    ws4.cell(row=rr, column=3, value="☐").alignment = CEN  # 待勾选
    ws4.cell(row=rr, column=4, value="").font = NORM
    ws4.cell(row=rr, column=5, value=note).font = SMALL
    ws4.cell(row=rr, column=5).alignment = LEFT
box(ws4, 3, 1, 3 + len(CHECKS), 5)
ws4.column_dimensions["A"].width = 5
ws4.column_dimensions["B"].width = 60
ws4.column_dimensions["C"].width = 10
ws4.column_dimensions["D"].width = 12
ws4.column_dimensions["E"].width = 18
ws4.freeze_panes = "A4"

wb.save(OUT)
print("OK ->", OUT)
print("输入行数:", len(INPUT_ROWS))
